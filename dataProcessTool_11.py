import os
import pandas as pd
from openpyxl import load_workbook
from datetime import datetime
import tkinter as tk
from tkinter import ttk
import re
from xlsxwriter import Workbook

# 声明全局变量
global file1, start_date, end_date, save_address, filtered_df, product_name
global result_label, file_path_entry, start_date_entry, end_date_entry, save_address_entry

start_date = ""
end_date = ""
# 定义全局变量
# 统计当前的所有日期
all_keys = set()
product_data = dict()
# 相对路径
AllFilePath = dict()
all_refund_keys = dict()
statistics_refund_info = dict()

# 文件路径
# file1 = "D:\\PyCharm_Project\\DataPackage\\2023年粮曲检验中心原料验收统计情况表.xlsx"
file1 = ""


# 1、将基础表中的活动表切分并将信息转换为csv

# 读取文件
def get_filename(path_filename):
    """
    获取文件所在文件夹路径、带拓展文件名、文件名、拓展名
    :param path_filename: 带拓展完整路径
    :return: 文件所在文件夹路径、带拓展文件名、文件名、拓展名
    """
    (filepath, tempfilename) = os.path.split(path_filename)
    (filename, extension) = os.path.splitext(tempfilename)
    return filepath, tempfilename, filename, extension


# xlsx转csv
def xlsx2csv(data_filenamepath):
    """
    转换 xlsx -> csv
    :param data_filenamepath: xlsx 文件路径
    :return: csv 文件路径
    """
    filepath, tempfilename, filename, extension = get_filename(data_filenamepath)
    print(filepath)
    wb = load_workbook(data_filenamepath, data_only=True)
    for sheet_name in wb.sheetnames:
        sheet = wb[sheet_name]
        df = pd.DataFrame(sheet.values)
        df[0] = pd.to_datetime(df[0], format='%Y/%m/%d', errors='coerce')
        df[0] = df[0].dt.strftime('%Y/%m/%d')
        csv_filename = os.path.join(filepath, f"{sheet_name}.csv")
        print("csv_filename", csv_filename)
        AllFilePath.update({sheet_name: csv_filename})
        df.to_csv(csv_filename, index=False, header=False)


# 定义关键字列表
daogu_reason = ["色泽", "气味", "重金属", "水分", "金属物质", "不完善粒", "杂质", "出糙率", "整精米率", "黄粒米", "谷外糙米", "精米霉变粒", "稻谷霉变粒"]
yumi_reason = ["不完善粒超扣价", "活虫超扣价", "异味", "生霉粒超扣价", "水份超扣价", "杂质"]

fenmai_reason = ["不完善粒超扣价", "赤霉病粒超扣价", "虫破粒超扣价", "堆包不规范", "活虫超扣价", "粮食污染", "异味", "异味", "色泽异常", "水分超扣价", "杂质超扣价"]
gengnuomi_reason = ["不完善粒超扣价", "活虫超扣价", "异味", "生霉粒超扣价", "水份超扣价", "杂质"]
statistic_data_yumi = {i: 0 for i in yumi_reason}
statistic_data_fenmai = {i: 0 for i in fenmai_reason}
statistic_data_gengnuomi = {i: 0 for i in gengnuomi_reason}
statistic_data_genggaoliang = {i: 0 for i in daogu_reason}


def validate_date_format(date_str):
    # 验证输入是否符合日期格式 (YYYY/MM/DD)
    if re.match(r'^\d{4}/\d{2}/\d{2}$', date_str) is not None:
        return True
    else:
        return False


# 当前步骤可以简化到filter_data按钮中，为了方便bug修改，暂时不合并
def pathtest():
    paths = file_path_entry.get()
    file1 = paths.replace("\\", "\\" + "\\")
    if paths:
        result_label.config(text="已加载路径")
    else:
        result_label.config(text="输入为空,请重新输入!")


# 数据切分main函数
def split_data():
    # 转换为csv文件
    xlsx2csv(file1)
    result_label.config(text="数据已切分")


# 主流程main函数
def filter_data():
    '''

    :return:
    '''
    start_date = start_date_entry.get()
    end_date = end_date_entry.get()

    save_address = save_address_entry.get()

    # 验证日期格式是否正确
    if not validate_date_format(start_date) or not validate_date_format(end_date):
        result_label.config(text="日期格式不正确，请重新输入")
        return

    for sheetName, csv_filename in AllFilePath.items():
        df_yumi = pd.read_csv(csv_filename, skip_blank_lines=True, skiprows=1)
        df_yumi = pd.DataFrame(df_yumi)
        # 去除空行空列
        df_yumi = df_yumi.dropna(axis=1, how='all')
        df_yumi = df_yumi.dropna(axis=0, how='all')
        # 根据不同的产品名,使用不同的退货原因(之后修改)
        if "玉米" in sheetName:
            batchProcessing(df_yumi, yumi_reason)
        elif "粉麦" in sheetName:
            batchProcessing(df_yumi, fenmai_reason)
        elif "糯米" in sheetName:
            batchProcessing(df_yumi, gengnuomi_reason)
        elif "高粱" in sheetName:
            batchProcessing(df_yumi, daogu_reason)
        else:
            batchProcessing(df_yumi, daogu_reason)

    date_objects = [datetime.strptime(str(date), '%Y/%m/%d') for date in all_keys]
    sorted_dates = sorted(date_objects)
    # 打印排序后的日期列表
    all_keys1 = list()
    for date in sorted_dates:
        all_keys1.append(date.strftime('%Y/%m/%d'))

    df = pd.DataFrame({'日期': all_keys1})
    # 遍历每个品种的数据字典，将数据填写到表格中
    for product_name, product_data_dict in product_data.items():
        product_column = [product_data_dict[date] if date in product_data_dict else 0 for date in all_keys1]
        df[product_name] = product_column
    # 将NaN值替换为0（如果需要）
    df = df.fillna(0)
    # 打印最终的数据框
    # print(df)
    # 使用 Pandas 进行日期筛选

    filtered_df = df[(df['日期'] >= start_date) & (df['日期'] <= end_date)]
    # 将数据框写入到 xlsx 文件
    print("filtered_df")
    print(filtered_df)
    filtered_df.to_excel(f'{save_address}\\日统计表.xlsx', index=False, engine='openpyxl')
    result_label.config(text="已筛选数据")
    # 添加一个合计列
    refundDetail2xlsx()


# 最终数据输出
def refundDetail2xlsx():
    writer = pd.ExcelWriter(f'{save_address}\\原料退货信息周统计表.xlsx', engine='xlsxwriter')
    combined_df = pd.DataFrame(columns=['项目', '退货数量(kg)', '退货率(%)', '入库合格总量(kg)', '验收总数(kg)'])
    product_info = {}
    workbook = writer.book
    worksheet = workbook.add_worksheet("Combined_Products")
    # 设置格式
    header_format = workbook.add_format({
        'bold': True,
        'align': 'center',
        'valign': 'vcenter'
    })
    center_align_format = workbook.add_format({
        'align': 'center',
        'valign': 'vcenter'
    })
    worksheet.set_default_row(20)
    worksheet.set_column('A:Z', 18, center_align_format)
    worksheet.set_column('E:E', 21)
    top_header_format = workbook.add_format({
        'align': 'center',
        'valign': 'vcenter',
        'font_size': 16
    })
    worksheet.merge_range('A1:E1', "原料退货信息周统计表", top_header_format)

    current_row = 1

    for product_name, product_data_dict in statistics_refund_info.items():
        pd.options.display.float_format = '{:.2f}'.format
        df = pd.DataFrame.from_dict(product_data_dict, orient='index', columns=['退货数量(kg)'])
        df['项目'] = df.index
        df = df[['项目', '退货数量(kg)']]

        df['入库合格总量(kg)'] = filtered_df["玉米"].sum()
        df['验收总数(kg)'] = df['退货数量(kg)'].sum() + df['入库合格总量(kg)']
        df['退货率(%)'] = (df['退货数量(kg)'] / df['验收总数(kg)']) * 100

        # # Add a "累计" row for the product
        # df = df.append({'项目': '累   计', '退货数量(kg)': sum(df['退货数量(kg)']),
        #                 '入库合格总量(kg)': '', '验收总数(kg)': '',
        #                 '退货率(%)': sum(df['退货率(%)'])}, ignore_index=True)
        # Create a "累   计" DataFrame with the desired values
        cumulative_data = pd.DataFrame({
            '项目': ['累   计'],
            '退货数量(kg)': [sum(df['退货数量(kg)'])],
            '入库合格总量(kg)': [''],
            '验收总数(kg)': [''],
            '退货率(%)': [sum(df['退货率(%)'])]
        })

        df = pd.concat([df, cumulative_data], ignore_index=True)
        df = df[['项目', '退货数量(kg)', '退货率(%)', '入库合格总量(kg)', '验收总数(kg)']]
        df = df.fillna('')

        product_info[product_name] = {
            'unit': '单位：粮曲检验中心',
            'product_name': '品名：',
            'product_name_1': product_name,
            'return_date': '退货日期: ',
            'return_date_1': '' + start_date + '-' + end_date
        }

        worksheet.write(current_row, 0, product_info[product_name]['unit'])
        worksheet.write(current_row, 1, product_info[product_name]['product_name'])
        worksheet.write(current_row, 2, product_info[product_name]['product_name_1'])
        worksheet.write(current_row, 3, product_info[product_name]['return_date'])
        worksheet.write(current_row, 4, product_info[product_name]['return_date_1'])
        current_row += 1

        for col_num, value in enumerate(combined_df.columns.values):
            worksheet.write(current_row, col_num, value, header_format)
        current_row += 1

        for row, data in df.iterrows():
            for col_num, value in enumerate(data):
                worksheet.write(current_row + row, col_num, value)

        current_row += len(df) + 2

    workbook.close()
    print("Excel表格已生成")


# 数据计算方法
def batchProcessing(df_yumi, daogu_reason):
    """
    :param df_yumi:
    :param daogu_reason:
    :return:
    """
    """
        验收情况统计
    """
    # 每日入库数量统计表
    everyday_input_totals = dict()
    statistic_data_yumi = {i: 0 for i in daogu_reason}
    # 遍历DataFrame的每一行
    for index, row_info in df_yumi.iterrows():
        if pd.isna(row_info).all():  # 去除所有空行
            continue
        text = row_info["退货原因"]  # 获取当前行的文本列

        # 检查是否包含NaN
        if pd.notna(text):
            # product_name = row_info["验收品种"]
            # 对每个关键字进行检查
            for keyword in daogu_reason:
                if keyword in text:
                    # supplier = row_info["供货单位"]
                    # descr = row_info["备注"]
                    statistic_data_yumi[keyword] += row_info['退货数量']
                    # 一种原因（关键词）只计算一次，避免关键词重复
                    break
            # 在填写了原因的情况下，如果关键字（原因）遍历完毕，仍不能定位备注原因的，单独记录
            else:
                if text in statistic_data_yumi:
                    statistic_data_yumi[text] += row_info['退货数量']
                    print("-------------------------1-----------------------------")
                else:
                    statistic_data_yumi[text] = 0
                    statistic_data_yumi[text] += row_info['退货数量']
                    print("-------------------------2-----------------------------")
        else:
            # 如果当前的退货原因为空，那么当前产品一定是入库了
            # print("退货原因为空，当前商品已入库！")
            #     计算每日的收获量
            pass
            # 将当前的日期改为日期格式或文本格式，否则数据只是一个flota，可以在表格数据处理的阶段将这些东西转换
            # cur_date = datetime.strptime(row_info["验收日期"], '%m.%d')
        # cur_date = row_info["验收日期"]
        cur_date = row_info[0]
        # cur_date = datetime.strptime(str(cur_date), '%Y/%m/%d').strftime('%m.%d')
        product_name = row_info["验收品种"]
        supplier = row_info["供货单位"]
        descr = row_info["备注"]
        acceptance_quantity = row_info["验收数量"]
        quantity_returned = row_info["退货数量"]
        if quantity_returned is None:
            print("测试当前是否全部遍历到(如果正确，结果为0)，退货数量为：", quantity_returned)
        storage_quantity = acceptance_quantity - quantity_returned
        if cur_date in everyday_input_totals:
            everyday_input_totals[cur_date] += storage_quantity
        else:
            everyday_input_totals[cur_date] = 0
            everyday_input_totals[cur_date] += storage_quantity
        # print(row_info)
    # print("smeel_row_count:",smell_row_count)

    # 日期：入库数量
    print("everyday_input_totals", everyday_input_totals)
    all_keys.update(everyday_input_totals.keys())
    product_data.update({product_name: everyday_input_totals})
    # print("product_data",product_data)
    # 产品名字：{原因：退货数量(kg),原因2：退货数量2(kg)}
    print("statistic_data_yumi", statistic_data_yumi)
    statistics_refund_info.update({product_name: statistic_data_yumi})
    print(statistics_refund_info)


"""
    信息写入表格，
    将csv表格转换为xlsx
"""


def center_window(root):
    # 获取屏幕宽度和高度
    screen_width = root.winfo_screenwidth()
    screen_height = root.winfo_screenheight()

    # 计算窗口的宽度和高度（这里使用屏幕大小的一半）
    window_width = screen_width // 2
    window_height = screen_height // 2

    # 计算窗口在屏幕中央的位置
    x = (screen_width - window_width) // 2
    y = (screen_height - window_height) // 2

    # 使用geometry方法设置窗口的位置和尺寸
    root.geometry(f"{window_width}x{window_height}+{x}+{y}")


def bindStartWindow():
    # 创建主窗口
    root = tk.Tk()
    root.title("表格一体化处理工具")
    # 调用center_window函数将窗口居中显示
    center_window(root)

    # 文件路径
    file_path_label = ttk.Label(root, text="输入文件路径:")
    file_path_label.pack()

    file_path_entry = ttk.Entry(root)
    file_path_entry.pack()

    # 文件路径测试
    pathtest_button = ttk.Button(root, text="加载路径", command=pathtest)
    pathtest_button.pack()

    # 表格切分
    split_button = ttk.Button(root, text="切分基础表格", command=split_data)
    split_button.pack()

    # 创建标签和输入框用于输入开始日期
    start_date_label = ttk.Label(root, text="开始日期:")
    start_date_label.pack()
    start_date_entry = ttk.Entry(root)
    start_date_entry.pack()

    # 创建标签和输入框用于输入结束日期
    end_date_label = ttk.Label(root, text="结束日期:")
    end_date_label.pack()
    end_date_entry = ttk.Entry(root)
    end_date_entry.pack()

    # 创建标签和输入框用于输入结束日期
    save_address_label = ttk.Label(root, text="保存地址:")
    save_address_label.pack()
    save_address_entry = ttk.Entry(root)
    save_address_entry.pack()

    # 创建按钮用于执行筛选操作
    filter_button = ttk.Button(root, text="筛选并保存数据", command=filter_data)
    filter_button.pack()

    # 创建标签用于显示筛选结果

    result_label = ttk.Label(root, text="")
    result_label.pack()

    # 运行主循环
    root.mainloop()


if __name__ == '__main__':
    bindStartWindow()
